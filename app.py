import os
import shutil
import unicodedata
import re
import pdfplumber
import tempfile
import json
import subprocess
from flask import Flask, jsonify, send_from_directory, request, send_file
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from io import BytesIO
import zipfile

# Configurar o caminho correto para os arquivos estáticos
static_folder = os.path.join(os.path.dirname(__file__), 'dist', 'public')
if not os.path.exists(static_folder):
    static_folder = os.path.join(os.path.dirname(__file__), 'dist')

app = Flask(__name__, static_folder=static_folder, static_url_path='')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max

# Configurações de Caminhos
BASE_PATH = os.getenv("BASE_PATH", r"G:\Wallpaper\FEDCORP_PROCESSADOR")
ENTRADA_PATH = os.path.join(BASE_PATH, "ENTRADA")
GERADAS_PATH = os.path.join(BASE_PATH, "REMESSAS_GERADAS")
NAO_PROCESSADOS_PATH = os.path.join(BASE_PATH, "NAO_PROCESSADOS")

# Armazenamento local no Render
PASTA_DOCS_PATH = os.path.join("/tmp", "fedcorp_docs")
os.makedirs(PASTA_DOCS_PATH, exist_ok=True)

# Tentar usar pasta local se disponível
if os.path.exists(r"G:\Wallpaper\FEDCORP_PROCESSADOR"):
    PASTA_DOCS_PATH = os.path.join(BASE_PATH, "DOCUMENTOS_ANEXADOS")

# Possíveis caminhos para o arquivo de condominios
POSSIBLE_PATHS = [
    os.path.join(BASE_PATH, "BASE", "DADOS_CONDOMINIOS.xlsx"),
    os.path.join(os.path.dirname(__file__), "BASE", "DADOS_CONDOMINIOS.xlsx"),
    os.path.join(os.path.dirname(__file__), "DADOS_CONDOMINIOS.xlsx"),
    "/app/BASE/DADOS_CONDOMINIOS.xlsx",
    "BASE/DADOS_CONDOMINIOS.xlsx",
]

# Pasta temporária para uploads
TEMP_UPLOAD_PATH = os.path.join(tempfile.gettempdir(), "fedcorp_uploads")
os.makedirs(TEMP_UPLOAD_PATH, exist_ok=True)

# ============================================================================
# MÓDULO FEDCORP - Dados Fixos para Seguro de Vida
# ============================================================================
FEDCORP_CONFIG = {
    "cnpj_admin": "26231209000150",
    "nome_admin": "GW ADMINISTRADORA DE CONDOMINIOS LTDA",
    "fornecedor_cnpj": "35315360000167",
    "fornecedor_nome": "FEDCORP ADMINISTRADORA DE BENEFICIOS LTDA",
    "cod_fornecedor_erp": "24196",
    "cod_produto_erp": "SEGUROVIDA",
    "desc_produto_erp": "",
    "tipo": "BOLETO"
}

# ============================================================================
# MÓDULO CONDOMED - Dados Fixos para Medicina e Seg. do Trabalho
# ============================================================================
CONDOMED_CONFIG = {
    "cnpj_admin": "26231209000150",
    "nome_admin": "GW ADMINISTRADORA DE CONDOMINIOS LTDA",
    "fornecedor_cnpj": "27892999000187",
    "fornecedor_nome": "CONDOMED RIO SEGURANCA E MEDICINA DO TRABALHO LTDA",
    "cod_fornecedor_erp": "24367",
    "cod_produto_erp": "MST",
    "desc_produto_erp": "",
    "tipo": "NFS-E"
}

# Cache de condominios em memória
CONDOMINIOS_CACHE = None
CACHE_TIMESTAMP = None

def remover_acentos(texto):
    if not texto: return ""
    return unicodedata.normalize("NFKD", str(texto)).encode("ASCII", "ignore").decode("ASCII")

def fixo(texto, tamanho):
    """Preenche texto com espaços em branco até o tamanho especificado"""
    return str(texto).ljust(tamanho)[:tamanho]

def extrair_cnpj_do_nome_arquivo(nome_arquivo):
    match = re.search(r"\d{14}", nome_arquivo)
    return match.group() if match else None

def detectar_tipo_documento(pdf_path):
    """Detecta se é boleto FEDCORP ou NFS-e CONDOMED"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            texto_completo = ""
            for pagina in pdf.pages:
                t = pagina.extract_text()
                if t: texto_completo += t + "\n"
            
            # Detectar NFS-e
            if "NFS-e" in texto_completo or "Nota Fiscal de Serviços" in texto_completo:
                if "CONDOMED" in texto_completo or "27.892.999" in texto_completo:
                    return "CONDOMED"
            
            # Detectar boleto FEDCORP
            if "FEDCORP" in texto_completo or "35.315.360" in texto_completo:
                return "FEDCORP"
            
            # Se tiver linha digitável, é boleto
            regex_linha = r"\d{5}[\.\s]?\d{5}[\.\s]?\d{5}[\.\s]?\d{6}[\.\s]?\d{5}[\.\s]?\d{6}[\.\s]?\d[\.\s]?\d{14}"
            if re.search(regex_linha, texto_completo):
                return "FEDCORP"
            
            # Se tiver NFS-e, é CONDOMED
            if "DANFSe" in texto_completo or "Chave de Acesso da NFS-e" in texto_completo:
                return "CONDOMED"
    
    except Exception as e:
        print(f"Erro ao detectar tipo: {e}")
    
    return None

def extrair_dados_boleto(pdf_path):
    """Extrai dados de boleto FEDCORP"""
    dados = {
        "linha_digitavel": None,
        "numero_nota": None,
        "vencimento": None,
        "valor": None,
        "numero_nfse": None
    }
    try:
        with pdfplumber.open(pdf_path) as pdf:
            texto_completo = ""
            for pagina in pdf.pages:
                t = pagina.extract_text()
                if t: texto_completo += t + "\n"
            
            # Extrair linha digitável
            regex_linha = r"\d{5}[\.\s]?\d{5}[\.\s]?\d{5}[\.\s]?\d{6}[\.\s]?\d{5}[\.\s]?\d{6}[\.\s]?\d[\.\s]?\d{14}"
            match_linha = re.search(regex_linha, texto_completo)
            if match_linha:
                dados["linha_digitavel"] = re.sub(r"\D", "", match_linha.group())
            
            # Extrair número da nota
            match_nota = re.search(r"(?:FATURA|NOTA|DOC|Nº|NUMERO)[:\s]+(\d+)", texto_completo, re.IGNORECASE)
            if match_nota:
                dados["numero_nota"] = match_nota.group(1)
            
            # Extrair vencimento
            match_venc = re.search(r"Vencimento\s+(\d{2}/\d{2}/\d{4})", texto_completo, re.IGNORECASE)
            if match_venc:
                dados["vencimento"] = match_venc.group(1)
            else:
                match_venc = re.search(r"ATE O VENCIMENTO\s+(\d{2}/\d{2}/\d{4})", texto_completo)
                if match_venc:
                    dados["vencimento"] = match_venc.group(1)
            
            # Extrair valor
            match_valor = re.search(r"VALOR TOTAL:?\s*R\$\s*([\d\.,]+)", texto_completo, re.IGNORECASE)
            if match_valor:
                dados["valor"] = match_valor.group(1).replace(".", "").replace(",", ".")
    except Exception as e:
        print(f"Erro ao extrair dados do boleto: {e}")
    
    return dados

def extrair_dados_nfse(pdf_path):
    """Extrai dados de NFS-e CONDOMED"""
    dados = {
        "linha_digitavel": None,
        "numero_nota": None,
        "vencimento": None,
        "valor": None,
        "numero_nfse": None,
        "cnpj_pagador": None
    }
    try:
        with pdfplumber.open(pdf_path) as pdf:
            texto_completo = ""
            for pagina in pdf.pages:
                t = pagina.extract_text()
                if t: texto_completo += t + "\n"
            
            # Extrair número da NFS-e
            # Padrão: "Número da NFS-e" seguido de números na próxima linha
            match_nfse = re.search(r"Número da NFS-e[^\n]*\n\s*(\d+)\s+\d{2}/\d{2}/\d{4}", texto_completo)
            if match_nfse:
                dados["numero_nfse"] = match_nfse.group(1)
            else:
                # Tenta padrão alternativo
                match_nfse = re.search(r"Número da NFS-e\s+(\d+)", texto_completo)
                if match_nfse:
                    dados["numero_nfse"] = match_nfse.group(1)
            
            # Extrair CNPJ do pagador (tomador do serviço)
            # Tenta primeiro o padrão formatado
            match_cnpj = re.search(r"TOMADOR DO SERVIÇO\s+CNPJ / CPF / NIF\s+(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})", texto_completo, re.DOTALL)
            if not match_cnpj:
                # Tenta padrão sem formatação
                match_cnpj = re.search(r"CNPJ/CPF:\s*(\d{2}\d{3}\d{3}\d{4}\d{2})", texto_completo)
            if not match_cnpj:
                # Tenta outro padrão
                match_cnpj = re.search(r"Pagador\s+.*?CNPJ/CPF:\s*(\d+)", texto_completo, re.DOTALL)
            
            if match_cnpj:
                cnpj_raw = match_cnpj.group(1)
                # Se estiver formatado, remover formatação
                if "." in cnpj_raw or "-" in cnpj_raw or "/" in cnpj_raw:
                    dados["cnpj_pagador"] = cnpj_raw
                else:
                    # Se não estiver formatado, formatar
                    cnpj_clean = cnpj_raw.replace(".", "").replace("-", "").replace("/", "")
                    if len(cnpj_clean) == 14:
                        dados["cnpj_pagador"] = f"{cnpj_clean[0:2]}.{cnpj_clean[2:5]}.{cnpj_clean[5:8]}/{cnpj_clean[8:12]}-{cnpj_clean[12:14]}"
            
            # Extrair valor total da NFS-e
            match_valor = re.search(r"Valor Líquido da NFS-e\s+R\$\s*([\d\.,]+)", texto_completo)
            if match_valor:
                dados["valor"] = match_valor.group(1).replace(".", "").replace(",", ".")
            else:
                # Tenta outro padrão
                match_valor = re.search(r"VALOR TOTAL:?\s*R\$\s*([\d\.,]+)", texto_completo, re.IGNORECASE)
                if match_valor:
                    dados["valor"] = match_valor.group(1).replace(".", "").replace(",", ".")
            
            # Extrair vencimento
            match_venc = re.search(r"Vencimento\s+(\d{2}/\d{2}/\d{4})", texto_completo, re.IGNORECASE)
            if match_venc:
                dados["vencimento"] = match_venc.group(1)
            else:
                match_venc = re.search(r"ATE O VENCIMENTO\s+(\d{2}/\d{2}/\d{4})", texto_completo)
                if match_venc:
                    dados["vencimento"] = match_venc.group(1)
            
            # Extrair linha digitável (se houver boleto integrado)
            regex_linha = r"\d{5}[\.\s]?\d{5}[\.\s]?\d{5}[\.\s]?\d{6}[\.\s]?\d{5}[\.\s]?\d{6}[\.\s]?\d[\.\s]?\d{14}"
            match_linha = re.search(regex_linha, texto_completo)
            if match_linha:
                dados["linha_digitavel"] = re.sub(r"\D", "", match_linha.group())
    
    except Exception as e:
        print(f"Erro ao extrair dados da NFS-e: {e}")
    
    return dados

def linha_digitavel_para_codigo_barras(linha):
    """Converte linha digitável para código de barras"""
    linha = re.sub(r"\D", "", linha)
    
    if len(linha) == 50:
        linha = linha[:47]
    
    if len(linha) != 47:
        return None
    
    try:
        banco = linha[0:3]
        moeda = linha[3:4]
        campo1 = linha[4:9]
        campo2 = linha[10:20]
        campo3 = linha[21:31]
        dv_geral = linha[32:33]
        fator_venc_valor = linha[33:47]
        
        codigo_barras = banco + moeda + dv_geral + fator_venc_valor + campo1 + campo2 + campo3
        return codigo_barras
    except:
        return None

def formatar_valor_ahreas(valor_float):
    """Formata valor para 12 posições com vírgula"""
    return f"{valor_float:012.2f}".replace(".", ",")

def carregar_condominios():
    """Carrega condominios de arquivo Excel"""
    global CONDOMINIOS_CACHE, CACHE_TIMESTAMP
    
    if CONDOMINIOS_CACHE is not None:
        if CACHE_TIMESTAMP and (datetime.now() - CACHE_TIMESTAMP).seconds < 300:
            return CONDOMINIOS_CACHE
    
    condominios = {}
    arquivo_encontrado = None
    
    for caminho in POSSIBLE_PATHS:
        if os.path.exists(caminho):
            arquivo_encontrado = caminho
            print(f"✅ Arquivo de condominios encontrado em: {caminho}")
            break
    
    if arquivo_encontrado:
        try:
            wb = load_workbook(arquivo_encontrado)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[3]:
                    cnpj = str(row[3]).strip()
                    cnpj = cnpj.replace(".", "").replace("-", "").replace("/", "")
                    if len(cnpj) >= 14:
                        cnpj = cnpj[:14]
                    
                    condominios[cnpj] = {
                        "nome": str(row[1]) if row[1] else "",
                        "codigo": str(int(row[0])).zfill(4) if row[0] else "0000"
                    }
            print(f"✅ Carregados {len(condominios)} condominios")
        except Exception as e:
            print(f"⚠️ Erro ao carregar: {e}")
    else:
        print("⚠️ Arquivo não encontrado, usando dados embutidos")
    
    if not condominios:
        condominios = {
            "65169906000180": {
                "nome": "CONDOMINIO EDIFICIO GROPIUS",
                "codigo": "0762"
            },
            "25530528000101": {
                "nome": "CONDOMINIO DO EDIFICIO PARQUE SAO BENTO",
                "codigo": "0009"
            }
        }
    
    CONDOMINIOS_CACHE = condominios
    CACHE_TIMESTAMP = datetime.now()
    
    return condominios

def processar_arquivo(nome_arquivo, caminho_entrada=None):
    """Processa um arquivo PDF (boleto ou NFS-e)"""
    resultado = {
        "arquivo": nome_arquivo,
        "status": "erro",
        "mensagem": "",
        "dados": None,
        "tipo": None
    }
    
    try:
        if caminho_entrada is None:
            caminho_entrada = os.path.join(ENTRADA_PATH, nome_arquivo)
        
        if not os.path.exists(caminho_entrada):
            resultado["mensagem"] = f"Arquivo não encontrado"
            return resultado
        
        # Detectar tipo de documento
        tipo_doc = detectar_tipo_documento(caminho_entrada)
        resultado["tipo"] = tipo_doc
        
        if tipo_doc == "CONDOMED":
            return processar_nfse(nome_arquivo, caminho_entrada)
        else:
            return processar_boleto(nome_arquivo, caminho_entrada)
    
    except Exception as e:
        resultado["mensagem"] = f"Erro ao processar: {str(e)}"
        import traceback
        print(traceback.format_exc())
    
    return resultado

def processar_boleto(nome_arquivo, caminho_entrada):
    """Processa boleto FEDCORP"""
    resultado = {
        "arquivo": nome_arquivo,
        "status": "erro",
        "mensagem": "",
        "dados": None,
        "tipo": "FEDCORP"
    }
    
    try:
        # Extrair CNPJ
        cnpj_condominio = extrair_cnpj_do_nome_arquivo(nome_arquivo)
        if not cnpj_condominio:
            resultado["mensagem"] = "CNPJ não encontrado no nome"
            return resultado
        
        # Carregar condominios
        condominios = carregar_condominios()
        
        if cnpj_condominio not in condominios:
            cnpj_limpo = cnpj_condominio.replace(".", "").replace("-", "").replace("/", "")
            if len(cnpj_limpo) >= 14:
                cnpj_limpo = cnpj_limpo[:14]
            if cnpj_limpo not in condominios:
                resultado["mensagem"] = f"Condomínio com CNPJ {cnpj_condominio} não cadastrado"
                return resultado
            cnpj_condominio = cnpj_limpo
        
        cond_info = condominios[cnpj_condominio]
        
        # Extrair dados do PDF
        dados_pdf = extrair_dados_boleto(caminho_entrada)
        if not dados_pdf["linha_digitavel"]:
            resultado["mensagem"] = "Não foi possível extrair a linha digitável"
            return resultado
        
        # Converter para código de barras
        codigo_barras = linha_digitavel_para_codigo_barras(dados_pdf["linha_digitavel"])
        if not codigo_barras:
            resultado["mensagem"] = "Código de barras inválido"
            return resultado
        
        # Preparar dados
        agora = datetime.now()
        vencimento = dados_pdf["vencimento"] if dados_pdf["vencimento"] else agora.strftime("%d/%m/%Y")
        data_emissao = agora.strftime("%d/%m/%Y")
        valor_float = float(dados_pdf["valor"]) if dados_pdf["valor"] else 0.0
        valor_formatado = formatar_valor_ahreas(valor_float)
        cod_cond_erp = cond_info["codigo"].zfill(4)
        nome_cond_erp = fixo(remover_acentos(cond_info["nome"]).upper(), 50)
        
        # Copiar para pasta local
        ano_atual = agora.strftime("%Y")
        mes_atual = agora.strftime("%m")
        pasta_destino_docs = os.path.join(PASTA_DOCS_PATH, ano_atual, mes_atual)
        os.makedirs(pasta_destino_docs, exist_ok=True)
        caminho_pdf_destino = os.path.join(pasta_destino_docs, nome_arquivo)
        try:
            shutil.copy(caminho_entrada, caminho_pdf_destino)
        except Exception as e:
            print(f"Aviso: Não foi possível copiar para pasta local: {e}")
        
        # Gerar URL local/remota
        if os.getenv('RENDER'):
            render_url = os.getenv('RENDER_EXTERNAL_URL', 'https://fedcorp-erp-dashboard.onrender.com')
            url_local = f"{render_url}/docs/{ano_atual}/{mes_atual}/{nome_arquivo}"
        else:
            url_local = f"http://localhost:5000/docs/{ano_atual}/{mes_atual}/{nome_arquivo}"
        
        # Retornar dados
        resultado["status"] = "sucesso"
        resultado["mensagem"] = "Boleto processado com sucesso"
        resultado["dados"] = {
            "cnpj": cnpj_condominio,
            "cod_cond": cod_cond_erp,
            "nome_cond": nome_cond_erp,
            "vencimento": vencimento,
            "data_emissao": data_emissao,
            "valor_formatado": valor_formatado,
            "valor_float": valor_float,
            "codigo_barras": codigo_barras,
            "nome_arquivo": nome_arquivo,
            "caminho_pdf": caminho_pdf_destino,
            "url_local": url_local,
            "ano": ano_atual,
            "mes": mes_atual,
            "numero_nfse": None,
            "tipo_fornecedor": "FEDCORP"
        }
    
    except Exception as e:
        resultado["mensagem"] = f"Erro ao processar boleto: {str(e)}"
        import traceback
        print(traceback.format_exc())
    
    return resultado

def processar_nfse(nome_arquivo, caminho_entrada):
    """Processa NFS-e CONDOMED"""
    resultado = {
        "arquivo": nome_arquivo,
        "status": "erro",
        "mensagem": "",
        "dados": None,
        "tipo": "CONDOMED"
    }
    
    try:
        # Extrair dados da NFS-e
        dados_pdf = extrair_dados_nfse(caminho_entrada)
        
        if not dados_pdf["cnpj_pagador"]:
            resultado["mensagem"] = "Não foi possível extrair CNPJ do pagador"
            return resultado
        
        # Limpar CNPJ
        cnpj_condominio = dados_pdf["cnpj_pagador"].replace(".", "").replace("-", "").replace("/", "")
        if len(cnpj_condominio) >= 14:
            cnpj_condominio = cnpj_condominio[:14]
        
        # Carregar condominios
        condominios = carregar_condominios()
        
        if cnpj_condominio not in condominios:
            resultado["mensagem"] = f"Condomínio com CNPJ {cnpj_condominio} não cadastrado"
            return resultado
        
        cond_info = condominios[cnpj_condominio]
        
        # Preparar dados
        agora = datetime.now()
        vencimento = dados_pdf["vencimento"] if dados_pdf["vencimento"] else agora.strftime("%d/%m/%Y")
        data_emissao = agora.strftime("%d/%m/%Y")
        valor_float = float(dados_pdf["valor"]) if dados_pdf["valor"] else 0.0
        valor_formatado = formatar_valor_ahreas(valor_float)
        cod_cond_erp = cond_info["codigo"].zfill(4)
        nome_cond_erp = fixo(remover_acentos(cond_info["nome"]).upper(), 50)
        
        # Converter linha digitável para código de barras (como em FEDCORP)
        if not dados_pdf["linha_digitavel"]:
            resultado["mensagem"] = "Não foi possível extrair a linha digitável do boleto"
            return resultado
        
        codigo_barras = linha_digitavel_para_codigo_barras(dados_pdf["linha_digitavel"])
        if not codigo_barras:
            resultado["mensagem"] = "Código de barras inválido"
            return resultado
        
        numero_nfse = dados_pdf["numero_nfse"] if dados_pdf["numero_nfse"] else None
        
        # Copiar para pasta local
        ano_atual = agora.strftime("%Y")
        mes_atual = agora.strftime("%m")
        pasta_destino_docs = os.path.join(PASTA_DOCS_PATH, ano_atual, mes_atual)
        os.makedirs(pasta_destino_docs, exist_ok=True)
        caminho_pdf_destino = os.path.join(pasta_destino_docs, nome_arquivo)
        try:
            shutil.copy(caminho_entrada, caminho_pdf_destino)
        except Exception as e:
            print(f"Aviso: Não foi possível copiar para pasta local: {e}")
        
        # Gerar URL local/remota
        if os.getenv('RENDER'):
            render_url = os.getenv('RENDER_EXTERNAL_URL', 'https://fedcorp-erp-dashboard.onrender.com')
            url_local = f"{render_url}/docs/{ano_atual}/{mes_atual}/{nome_arquivo}"
        else:
            url_local = f"http://localhost:5000/docs/{ano_atual}/{mes_atual}/{nome_arquivo}"
        
        # Retornar dados
        resultado["status"] = "sucesso"
        resultado["mensagem"] = "NFS-e processada com sucesso"
        resultado["dados"] = {
            "cnpj": cnpj_condominio,
            "cod_cond": cod_cond_erp,
            "nome_cond": nome_cond_erp,
            "vencimento": vencimento,
            "data_emissao": data_emissao,
            "valor_formatado": valor_formatado,
            "valor_float": valor_float,
            "codigo_barras": codigo_barras,
            "nome_arquivo": nome_arquivo,
            "caminho_pdf": caminho_pdf_destino,
            "url_local": url_local,
            "ano": ano_atual,
            "mes": mes_atual,
            "numero_nfse": dados_pdf["numero_nfse"],
            "tipo_fornecedor": "CONDOMED"
        }
    
    except Exception as e:
        resultado["mensagem"] = f"Erro ao processar NFS-e: {str(e)}"
        import traceback
        print(traceback.format_exc())
    
    return resultado

def gerar_remessa_lote(lista_dados, competencia=None):
    """Gera remessa única com múltiplos registros (boletos e NFS-e)"""
    if not lista_dados:
        return None
    
    agora = datetime.now()
    if not competencia:
        competencia = agora.strftime("%m%Y")
    
    linhas = []
    
    # Determinar fornecedor baseado no primeiro item
    tipo_primeiro = lista_dados[0].get("tipo_fornecedor", "FEDCORP")
    
    if tipo_primeiro == "CONDOMED":
        config = CONDOMED_CONFIG
    else:
        config = FEDCORP_CONFIG
    
    # REGISTRO 0 - HEADER
    header = (
        "0" +
        config["fornecedor_cnpj"].zfill(14) +
        fixo(remover_acentos(config["fornecedor_nome"]).upper(), 60) +
        config["cnpj_admin"].zfill(14) +
        fixo(remover_acentos(config["nome_admin"]).upper(), 60) +
        competencia +
        " " * 241 +
        "0001"
    )
    linhas.append(fixo(header, 400))
    
    # REGISTROS 1, 2 e 3 para cada documento
    sequencial = 2
    for dados in lista_dados:
        # REGISTRO 1 - Conforme layout de importação (400 chars)
        # Campos conforme especificação do layout
        registro_1 = (
            "1" +                                                   # Pos 001: Tipo de registro
            fixo(dados["cod_cond"], 4) +                           # Pos 002-005: Código do condomínio
            "0   " +                                                # Pos 006-009: Código do bloco (0 + 3 brancos)
            dados["cnpj"].zfill(14) +                              # Pos 010-023: CNPJ
            fixo(dados["nome_cond"], 50) +                         # Pos 024-073: Nome do condomínio
            fixo(dados["vencimento"], 10) +                        # Pos 074-083: Data vencimento (DD/MM/AAAA)
            fixo(dados["valor_formatado"], 12) +                   # Pos 084-095: Valor do título (líquido)
            fixo(dados["codigo_barras"], 44) +                     # Pos 096-139: Código de barras (44 chars!)
            fixo(dados["valor_formatado"], 12) +                   # Pos 140-151: Valor total da NF sem retenção
            "000000000,00" +                                        # Pos 152-163: IRRF
            "000000000,00" +                                        # Pos 164-175: ISS
            "000000000,00" +                                        # Pos 176-187: INSS
            "000000000,00" +                                        # Pos 188-199: CSSL/PIS/COFINS
            "000000000,00" +                                        # Pos 200-211: Descontos
            "N" +                                                   # Pos 212: Nota fiscal de venda (S/N)
            fixo(dados.get("data_emissao_nf", dados["data_emissao"]), 10) +  # Pos 213-222: Data emissão NF
            fixo(dados.get("numero_nfse", ""), 10) +               # Pos 223-232: Número da NF
            "     " +                                                # Pos 233-237: Série da NF
            "     " +                                                # Pos 238-242: Tipo da NF
            "000000000,00" +                                        # Pos 243-254: CSLL
            "000000000,00" +                                        # Pos 255-266: PIS
            "000000000,00" +                                        # Pos 267-278: COFINS
            " " * 118 +                                             # Pos 279-396: Uso Ahreas
            str(sequencial).zfill(4)                                # Pos 397-400: Sequencial
        )
        linhas.append(fixo(registro_1, 400))
        
        # REGISTRO 2 - Exatamente igual ao original FEDCORP
        registro_2 = (
            "2" +
            fixo(config["cod_produto_erp"], 10) +
            fixo(config["desc_produto_erp"], 60) +
            "000000000,00" +
            dados["valor_formatado"] +
            dados["valor_formatado"] +
            " " * 289 +
            str(sequencial).zfill(4)
        )
        linhas.append(fixo(registro_2, 400))
        
        # REGISTRO 3 - URL do PDF (sem o número NFS-e, que já está no Registro 1)
        url_pdf = dados.get("url_local", "")
        
        trailer_boleto = (
            "3" +
            "0001" +
            "0000000000" +
            fixo(url_pdf, 300) +
            " " * 81 +
            str(sequencial).zfill(4)
        )
        linhas.append(fixo(trailer_boleto, 400))
        
        sequencial += 1
    
    return "\n".join(linhas)

# Rotas da API
@app.route('/')
def index():
    index_path = os.path.join(app.static_folder, 'index.html')
    if os.path.exists(index_path):
        return send_from_directory(app.static_folder, 'index.html')
    else:
        return jsonify({
            "status": "error",
            "message": "index.html não encontrado",
            "static_folder": app.static_folder
        }), 404

@app.route('/api/pending-files', methods=['GET'])
def pending_files():
    try:
        os.makedirs(ENTRADA_PATH, exist_ok=True)
        arquivos = [f for f in os.listdir(ENTRADA_PATH) if f.lower().endswith('.pdf')]
        return jsonify({
            "total": len(arquivos),
            "arquivos": arquivos
        })
    except Exception as e:
        return jsonify({"erro": str(e)}), 500

@app.route('/api/upload', methods=['POST'])
def upload_files():
    try:
        if 'files' not in request.files:
            return jsonify({"erro": "Nenhum arquivo enviado"}), 400
        
        files = request.files.getlist('files')
        modo_lote = request.form.get('modo_lote', 'false').lower() == 'true'
        
        resultados = {
            "total": len(files),
            "sucesso": 0,
            "erros": 0,
            "modo_lote": modo_lote,
            "remessa": None,
            "detalhes": []
        }
        
        lista_dados_processados = []
        
        for file in files:
            if file and file.filename.lower().endswith('.pdf'):
                try:
                    filename = secure_filename(file.filename)
                    temp_path = os.path.join(TEMP_UPLOAD_PATH, filename)
                    file.save(temp_path)
                    
                    resultado = processar_arquivo(filename, temp_path)
                    resultados["detalhes"].append(resultado)
                    
                    if resultado["status"] == "sucesso":
                        resultados["sucesso"] += 1
                        if modo_lote:
                            lista_dados_processados.append(resultado["dados"])
                    else:
                        resultados["erros"] += 1
                    
                    try:
                        os.remove(temp_path)
                    except:
                        pass
                
                except Exception as e:
                    resultados["erros"] += 1
                    resultados["detalhes"].append({
                        "arquivo": file.filename,
                        "status": "erro",
                        "mensagem": str(e)
                    })
        
        # Se modo lote
        if modo_lote and lista_dados_processados:
            try:
                agora = datetime.now()
                competencia = agora.strftime("%m%Y")
                conteudo_remessa = gerar_remessa_lote(lista_dados_processados, competencia)
                
                # Determinar nome baseado no tipo de fornecedor
                tipo_fornecedor = lista_dados_processados[0].get("tipo_fornecedor", "FEDCORP")
                prefixo = "REMESSA_CONDOMED" if tipo_fornecedor == "CONDOMED" else "REMESSA_FEDCORP"
                nome_remessa = f"{prefixo}_LOTE_{agora.strftime('%Y%m%d%H%M%S')}.txt"
                caminho_remessa = os.path.join(GERADAS_PATH, nome_remessa)
                
                os.makedirs(GERADAS_PATH, exist_ok=True)
                with open(caminho_remessa, 'w', encoding='utf-8') as f:
                    f.write(conteudo_remessa)
                
                resultados["remessa"] = nome_remessa
            
            except Exception as e:
                resultados["erro_lote"] = str(e)
        
        return jsonify(resultados)
    except Exception as e:
        return jsonify({"erro": str(e)}), 500

@app.route('/api/download-remessas', methods=['GET'])
def download_remessas():
    try:
        os.makedirs(GERADAS_PATH, exist_ok=True)
        
        remessas = [f for f in os.listdir(GERADAS_PATH) if f.endswith('.txt')]
        
        if not remessas:
            return jsonify({"erro": "Nenhuma remessa disponível"}), 404
        
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for remessa in remessas:
                remessa_path = os.path.join(GERADAS_PATH, remessa)
                zip_file.write(remessa_path, arcname=remessa)
        
        zip_buffer.seek(0)
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'remessas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
        )
    except Exception as e:
        return jsonify({"erro": str(e)}), 500

@app.route('/docs/<ano>/<mes>/<arquivo>')
def servir_documento(ano, mes, arquivo):
    try:
        caminho = os.path.join(PASTA_DOCS_PATH, ano, mes, arquivo)
        if os.path.exists(caminho):
            return send_from_directory(os.path.dirname(caminho), arquivo)
        return jsonify({"erro": "Arquivo não encontrado"}), 404
    except Exception as e:
        return jsonify({"erro": str(e)}), 500

@app.route('/<path:filename>')
def static_files(filename):
    try:
        return send_from_directory(app.static_folder, filename)
    except:
        try:
            return send_from_directory(app.static_folder, 'index.html')
        except:
            return jsonify({"erro": "Arquivo não encontrado"}), 404

if __name__ == '__main__':
    print("🚀 Iniciando FEDCORP ERP Dashboard v34...")
    print(f"📁 BASE_PATH: {BASE_PATH}")
    print(f"📂 PASTA_DOCS_PATH: {PASTA_DOCS_PATH}")
    print("📦 Módulos: FEDCORP (Boletos) + CONDOMED (NFS-e)")
    
    # Pré-carregar condominios
    print("📝 Carregando dados de condominios...")
    condominios = carregar_condominios()
    print(f"✅ {len(condominios)} condominio(s) carregado(s)")
    
    app.run(host='0.0.0.0', port=int(os.getenv('PORT', 5000)), debug=False)
